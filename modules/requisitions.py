# utils/excel_importer.py
import streamlit as st
import pandas as pd
from datetime import datetime
from database import get_db_connection
from utils.pdf_generator import generate_requisition_pdf
from utils.helpers import get_status_badge
from utils.excel_importer import import_excel_data  # This is correct now
import io
import os
import tempfile
import re
from openpyxl import load_workbook

class ExcelRequisitionImporter:
    """Template-independent Excel importer that dynamically detects categories"""
    
    def __init__(self, file_path, user_id):
        self.file_path = file_path
        self.user_id = user_id
        self.wb = None
        self.ws = None
        self.merged_cells = {}
        self.current_category = "Site Work"
        self.transactions = []
        self.metadata = {
            'ref_no': None,
            'period_start': None,
            'period_end': None,
            'expense_paid_last': 0,
            'total_amount': 0
        }
        self.row_continuation = False
        self.last_transaction = None
        
    def load_workbook(self):
        """Load the Excel workbook"""
        self.wb = load_workbook(self.file_path, data_only=True)
        self.ws = self.wb.active
        self._detect_merged_cells()
        
    def _detect_merged_cells(self):
        """Detect all merged cells and store their values"""
        for merged_range in self.ws.merged_cells.ranges:
            # Get the top-left cell value
            min_row = merged_range.min_row
            min_col = merged_range.min_col
            value = self.ws.cell(row=min_row, column=min_col).value
            
            if value:
                # Store the merged range with its value
                self.merged_cells[min_row] = {
                    'value': str(value).strip(),
                    'range': merged_range,
                    'min_col': min_col,
                    'max_col': merged_range.max_col
                }
    
    def _clean_value(self, value):
        """Clean and convert cell values"""
        if value is None:
            return 0
        if isinstance(value, (int, float)):
            return value
        if isinstance(value, datetime):
            return value
        if isinstance(value, str):
            raw_text = value.strip()
            if not raw_text:
                return 0
            # Remove commas and currency symbols
            cleaned = re.sub(r'[^\d.,-]', '', raw_text)
            if cleaned:
                try:
                    # Handle comma as thousand separator
                    if ',' in cleaned and '.' in cleaned:
                        cleaned = cleaned.replace(',', '')
                    elif ',' in cleaned:
                        cleaned = cleaned.replace(',', '')
                    return float(cleaned)
                except:
                    return 0
        return 0
    
    def _is_transaction_row(self, row_data):
        """Check if a row contains transaction data"""
        # Check first column for serial number
        first_val = row_data[0] if row_data else ''
        if isinstance(first_val, (int, float)):
            if first_val > 0 and first_val.is_integer():
                return True
        
        # Check for pattern: number, text, number, number, number
        if len(row_data) >= 5:
            # Check if first value is a number (serial)
            if isinstance(row_data[0], (int, float)):
                if row_data[0] > 0:
                    return True
            # Check if second value has text
            if row_data[1] and isinstance(row_data[1], str):
                # Check if there are numbers in later columns
                if any(isinstance(x, (int, float)) and x > 0 for x in row_data[2:5]):
                    return True
        return False
    
    def _is_continuation_row(self, row_data):
        """Check if this is a continuation of the previous transaction"""
        if not self.last_transaction:
            return False
        
        # Check if first column is empty or not a number
        first_val = row_data[0] if row_data else ''
        if not first_val or (isinstance(first_val, str) and not first_val.strip()):
            # Check if there's descriptive text
            if len(row_data) > 1 and row_data[1] and isinstance(row_data[1], str):
                # Check if no amounts in the row
                if not any(isinstance(x, (int, float)) and x > 0 for x in row_data[2:5]):
                    return True
        return False
    
    def _is_total_row(self, row_data):
        """Check if this is a total row"""
        row_text = ' '.join(str(x) for x in row_data if x).lower()
        return 'total' in row_text
    
    def _is_category_row(self, row_data):
        """Check if this row is a category header"""
        row_text = ' '.join(str(x) for x in row_data if x).strip()
        if not row_text:
            return False
        
        # Check if it's a merged cell (category)
        for row_num, merged_info in self.merged_cells.items():
            if row_num in row_data:
                return True
        
        # Check for common category patterns
        category_indicators = ['materials', 'miscellaneous', 'administration', 'work', 'labour', 
                              'electrical', 'plumbing', 'tiles', 'furniture', 'steel', 'cement']
        for indicator in category_indicators:
            if indicator in row_text.lower() and len(row_text) < 50:
                return True
        
        return False
    
    def _extract_category_from_row(self, row_data):
        """Extract category name from a row"""
        for row_num, merged_info in self.merged_cells.items():
            if row_num in row_data:
                return merged_info['value']
        
        # If not a merged cell, try to detect from text
        row_text = ' '.join(str(x) for x in row_data if x).strip()
        if row_text:
            # Capitalize and clean
            category = re.sub(r'[^\w\s]', '', row_text).strip()
            return category.title()
        
        return None
    
    def _extract_metadata(self, row_data):
        """Extract metadata like Ref No, Dates, etc."""
        row_text = ' '.join(str(x) for x in row_data if x)
        
        # Extract Reference Number
        ref_match = re.search(r'Ref\s*No\.?\s*[:.]?\s*(\S+)', row_text, re.IGNORECASE)
        if ref_match and not self.metadata['ref_no']:
            self.metadata['ref_no'] = ref_match.group(1)
        
        # Extract Dates
        # Support both numeric and month-name range formats like 17 Jul to 23 Jul 2026
        date_matches = re.findall(r'(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})', row_text)
        if len(date_matches) >= 2:
            try:
                start = date_matches[0]
                end = date_matches[1]
                self.metadata['period_start'] = datetime.strptime(
                    f"{start[0]}-{start[1]}-{start[2]}", '%d-%m-%Y'
                )
                self.metadata['period_end'] = datetime.strptime(
                    f"{end[0]}-{end[1]}-{end[2]}", '%d-%m-%Y'
                )
            except:
                pass
        else:
            month_name_matches = re.search(
                r'(\d{1,2})\s*([A-Za-z]{3,9})\s*(?:to|-)\s*(\d{1,2})\s*([A-Za-z]{3,9})\s*(\d{4})',
                row_text,
                re.IGNORECASE,
            )
            if month_name_matches:
                try:
                    start_day = month_name_matches.group(1)
                    start_month = month_name_matches.group(2)
                    end_day = month_name_matches.group(3)
                    end_month = month_name_matches.group(4)
                    year = month_name_matches.group(5)
                    self.metadata['period_start'] = datetime.strptime(
                        f"{start_day} {start_month} {year}", '%d %b %Y'
                    )
                    self.metadata['period_end'] = datetime.strptime(
                        f"{end_day} {end_month} {year}", '%d %b %Y'
                    )
                except:
                    pass

        # Extract Expense Paid Last
        expense_match = re.search(r'Expense\s*paid\s*last.*?([\d,]+\.?[\d]*)', row_text, re.IGNORECASE)
        if expense_match:
            try:
                self.metadata['expense_paid_last'] = float(
                    expense_match.group(1).replace(',', '')
                )
            except:
                pass
    
    def _get_or_create_category(self, category_name):
        """Get category ID from database, create if doesn't exist"""
        if not category_name:
            return None
            
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Check if category exists
        cursor.execute("SELECT id FROM categories WHERE name = ?", (category_name,))
        result = cursor.fetchone()
        
        if result:
            category_id = result['id']
        else:
            # Create new category
            cursor.execute(
                "INSERT INTO categories (name, sort_order, status) VALUES (?, ?, ?)",
                (category_name, 999, 'Active')
            )
            category_id = cursor.lastrowid
            st.info(f"🆕 New category created: {category_name}")
        
        conn.commit()
        conn.close()
        return category_id
    
    def _process_transaction(self, row_data, row_num):
        """Process a transaction row"""
        if not self._is_transaction_row(row_data):
            return
        
        # Extract data based on position
        sr_no = int(row_data[0]) if isinstance(row_data[0], (int, float)) else 0
        particulars = str(row_data[1]).strip() if len(row_data) > 1 and row_data[1] else ''
        qty = self._clean_value(row_data[2]) if len(row_data) > 2 else 0
        rate = self._clean_value(row_data[3]) if len(row_data) > 3 else 0
        amount = self._clean_value(row_data[4]) if len(row_data) > 4 else 0
        remarks = str(row_data[5]).strip() if len(row_data) > 5 and row_data[5] else ''
        
        # If amount is 0 but qty and rate exist, calculate
        if amount == 0 and qty > 0 and rate > 0:
            amount = qty * rate
        
        # Store transaction
        transaction = {
            'sr_no': sr_no,
            'particulars': particulars,
            'qty': float(qty),
            'rate': float(rate),
            'amount': float(amount),
            'remarks': remarks,
            'category': self.current_category,
            'row_num': row_num
        }
        
        self.transactions.append(transaction)
        self.last_transaction = transaction
        
        # Update total amount
        self.metadata['total_amount'] += float(amount)
    
    def _process_continuation(self, row_data):
        """Process a continuation row (append to previous transaction)"""
        if not self.last_transaction:
            return
        
        # Extract continuation text
        continuation_text = ''
        if len(row_data) > 1 and row_data[1]:
            continuation_text = str(row_data[1]).strip()
        elif len(row_data) > 0 and row_data[0]:
            continuation_text = str(row_data[0]).strip()
        
        if continuation_text:
            # Append to previous transaction's particulars
            self.last_transaction['particulars'] += f"\n{continuation_text}"
            
            # If there are remarks, add them too
            if len(row_data) > 5 and row_data[5]:
                self.last_transaction['remarks'] += f" {str(row_data[5]).strip()}"
    
    def import_data(self):
        """Main import function"""
        self.load_workbook()
        
        # Process each row
        for row_num in range(1, self.ws.max_row + 1):
            row_data = []
            for col in range(1, self.ws.max_column + 1):
                cell_value = self.ws.cell(row=row_num, column=col).value
                row_data.append(cell_value)
            
            # Skip empty rows
            if not any(row_data):
                continue
            
            # Check for merged cells (categories)
            is_category = False
            for merged_row, merged_info in self.merged_cells.items():
                if row_num == merged_row:
                    self.current_category = merged_info['value']
                    is_category = True
                    break
            
            # Extract metadata from header rows
            self._extract_metadata(row_data)
            
            # Skip if it's a category row (already processed)
            if is_category:
                continue
            
            # Check if it's a total row
            if self._is_total_row(row_data):
                continue
            
            # Check if it's a continuation row
            if self._is_continuation_row(row_data):
                self._process_continuation(row_data)
                continue
            
            # Process transaction
            if self._is_transaction_row(row_data):
                self._process_transaction(row_data, row_num)
        
        # Save to database
        return self._save_to_database()
    
    def _save_to_database(self):
        """Save imported data to database"""
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            # Get or create default project
            cursor.execute("SELECT id FROM projects LIMIT 1")
            project = cursor.fetchone()
            
            if not project:
                cursor.execute(
                    "INSERT INTO projects (name, code, opening_balance, status) VALUES (?, ?, ?, ?)",
                    ('Default Project', 'DEF', 0, 'Active')
                )
                project_id = cursor.lastrowid
            else:
                project_id = project['id']
            
            # Get opening balance from last requisition
            cursor.execute('''
                SELECT closing_balance FROM requisitions 
                WHERE project_id = ? AND status IN ('APPROVED', 'VERIFIED')
                ORDER BY created_at DESC LIMIT 1
            ''', (project_id,))
            last_req = cursor.fetchone()
            opening_balance = last_req['closing_balance'] if last_req else 0
            
            # Create requisition
            ref_no = self.metadata['ref_no'] or f"REQ-{datetime.now().strftime('%Y%m%d-%H%M')}"
            if not self.metadata['period_start']:
                self.metadata['period_start'] = datetime.now()
            if not self.metadata['period_end']:
                self.metadata['period_end'] = self.metadata['period_start']

            cursor.execute('''
                INSERT INTO requisitions 
                (project_id, ref_no, period_start, period_end, opening_balance, 
                 closing_balance, expense_paid_last_req, total_amount, status, created_by_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                project_id,
                ref_no,
                self.metadata['period_start'],
                self.metadata['period_end'],
                opening_balance,
                opening_balance + self.metadata['total_amount'],
                self.metadata['expense_paid_last'],
                self.metadata['total_amount'],
                'DRAFT',
                self.user_id
            ))
            
            requisition_id = cursor.lastrowid
            
            # Process and save transactions
            category_cache = {}
            for idx, trans in enumerate(self.transactions, 1):
                # Get or create category
                category_name = trans['category']
                if category_name not in category_cache:
                    category_id = self._get_or_create_category(category_name)
                    category_cache[category_name] = category_id
                else:
                    category_id = category_cache[category_name]
                
                # Insert transaction
                cursor.execute('''
                    INSERT INTO transactions 
                    (requisition_id, category_id, master_item_id, particulars_raw, qty, unit, 
                     rate, amount, remarks, sr_no, is_lump_sum, entered_by_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    requisition_id,
                    category_id,
                    None,
                    trans['particulars'],
                    trans['qty'],
                    'nos',
                    trans['rate'],
                    trans['amount'],
                    trans['remarks'],
                    idx,
                    1 if trans['qty'] == 0 and trans['rate'] == 0 and trans['amount'] > 0 else 0,
                    self.user_id
                ))
            
            conn.commit()
            conn.close()
            
            return True, f"✅ Successfully imported requisition {ref_no} with {len(self.transactions)} items", len(self.transactions)
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return False, f"Error saving to database: {str(e)}", 0


def show_categories():
    """Categories management page"""
    st.markdown('<div class="section-header">🏷️ Categories</div>', unsafe_allow_html=True)
    
    # Add new category
    with st.expander("➕ Add New Category", expanded=False):
        with st.form("add_category_form"):
            category_name = st.text_input("Category Name*")
            submitted = st.form_submit_button("Add Category", use_container_width=True, type="primary")
            
            if submitted and category_name:
                try:
                    conn = get_db_connection()
                    cursor = conn.cursor()
                    
                    # Check if category exists
                    cursor.execute("SELECT id FROM categories WHERE name = ?", (category_name,))
                    existing = cursor.fetchone()
                    
                    if existing:
                        st.warning(f"Category '{category_name}' already exists!")
                    else:
                        cursor.execute(
                            "INSERT INTO categories (name, sort_order, status) VALUES (?, ?, ?)",
                            (category_name, 999, 'Active')
                        )
                        conn.commit()
                        st.success(f"✅ Category '{category_name}' added successfully!")
                        st.rerun()
                    
                    conn.close()
                except Exception as e:
                    st.error(f"Error adding category: {e}")
    
    # Display categories
    st.subheader("All Categories")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, name, sort_order, status, created_at FROM categories ORDER BY sort_order, name")
    categories = cursor.fetchall()
    conn.close()
    
    if categories:
        # Convert to DataFrame for display
        data = []
        for cat in categories:
            data.append({
                "ID": cat['id'],
                "Category Name": cat['name'],
                "Sort Order": cat['sort_order'],
                "Status": cat['status'],
                "Created": cat['created_at'][:10] if cat['created_at'] else ''
            })
        
        df = pd.DataFrame(data)
        st.dataframe(df, use_container_width=True, hide_index=True)
        
        # Delete category option
        st.subheader("Delete Category")
        category_options = {f"{cat['name']} (ID: {cat['id']})": cat['id'] for cat in categories}
        selected_category = st.selectbox("Select category to delete", list(category_options.keys()))
        
        if st.button("🗑️ Delete Selected Category", type="primary"):
            category_id = category_options[selected_category]
            
            # Check if category has transactions
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) as count FROM transactions WHERE category_id = ?", (category_id,))
            result = cursor.fetchone()
            conn.close()
            
            if result and result['count'] > 0:
                st.error(f"Cannot delete category with {result['count']} transactions. Please reassign transactions first.")
            else:
                try:
                    conn = get_db_connection()
                    cursor = conn.cursor()
                    cursor.execute("DELETE FROM categories WHERE id = ?", (category_id,))
                    conn.commit()
                    conn.close()
                    st.success("✅ Category deleted successfully!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error deleting category: {e}")
    else:
        st.info("No categories found. Add your first category!")

def show_requisitions():
    """Main requisitions page with smart Excel import"""
    user = st.session_state.auth["user"]

    # ============================================================
    # DELETE CONFIRMATION DIALOG - MUST BE FIRST
    # ============================================================
    if st.session_state.get("delete_requisition_id"):
        req_id = st.session_state.delete_requisition_id
        
        # Get requisition details for confirmation
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT ref_no, status FROM requisitions WHERE id = ?", (req_id,))
        req = cursor.fetchone()
        conn.close()
        
        if req:
            st.warning(f"⚠️ Are you sure you want to delete requisition {req['ref_no']}?")
            st.info(f"Status: {req['status']} | This action cannot be undone!")
            
            col1, col2, col3 = st.columns([1, 1, 2])
            with col1:
                if st.button("✅ Yes, Delete", type="primary"):
                    try:
                        conn = get_db_connection()
                        cursor = conn.cursor()
                        
                        # First delete all transactions
                        cursor.execute("DELETE FROM transactions WHERE requisition_id = ?", (req_id,))
                        # Then delete the requisition
                        cursor.execute("DELETE FROM requisitions WHERE id = ?", (req_id,))
                        
                        conn.commit()
                        conn.close()
                        
                        st.session_state.delete_requisition_id = None
                        st.success("✅ Requisition deleted successfully!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error deleting requisition: {e}")
            
            with col2:
                if st.button("❌ Cancel"):
                    st.session_state.delete_requisition_id = None
                    st.rerun()
            
            # Don't show the rest of the page while confirmation is shown
            return

    # Rest of your function continues here...
    if st.session_state.get("editing_requisition_id"):
        show_requisition_editor(st.session_state.editing_requisition_id)
        return

    st.markdown(
        '<div class="section-header">📄 Requisitions</div>',
        unsafe_allow_html=True,
    )

    # ------------------------------------------------------------------
    # Import from Excel section
    # ------------------------------------------------------------------
    with st.expander("📥 Import from Excel", expanded=False):
        st.markdown('<div class="upload-section">', unsafe_allow_html=True)

        st.info("""
        **💡 Smart Import Feature**
        
        - ✅ Automatically detects merged rows as categories
        - ✅ Creates new categories on the fly
        - ✅ Handles continuation rows (like "Nazam Chokidar Salary")
        - ✅ Extracts metadata automatically (Ref No, Dates, etc.)
        - ✅ Future-proof - works with any Excel format
        """)

        # Show format example
        with st.expander("📋 View Expected Format"):
            st.markdown("""
            The system can handle various Excel formats. Here's an example:

            REQUISITION FORM FOR SITE WORK DONE
            Ref No.18
            Date.17-Jul-2025         To         Date.23-Jul-2025
            Sr. # | Particulars                    | Qty  | Rate   | Amount
            1     | Sun room work                  | 1    | 40000  | 40000
            Total |                                |      |        | 257500
            Materials                              |      |        |
            1     | Blocks                         | 1000 | 72     | 72000
            2     | Stone                          | 3    | 4000   | 12000
            Miscellaneous & Administration         |      |        |
            1     | Staff food                     |      |        | 2100
            | Nazam Chokidar Salary Jun 2026 |      |        |
            2     | Bike petrol                    |      |        | 500

            **Supported Features:**
            - ✅ Merged cells as categories (any name works)
            - ✅ Continuation rows (no serial number)
            - ✅ Various number formats (with commas)
            - ✅ Dynamic category creation
            - ✅ No template restrictions
            """)

        # Download sample template
        if st.button(
            "📄 Download Sample Excel Template", use_container_width=True
        ):
            sample_data = [
                [
                    "REQUISITION FORM FOR SITE WORK DONE",
                    "",
                    "",
                    "",
                    "Ref No.18",
                    "",
                ],
                ["Date.17-Jul-2025", "To", "Date.23-Jul-2025", "", "", ""],
                [
                    "Sr. #",
                    "Particulars",
                    "Qty",
                    "Market Rate",
                    "Amount",
                    "Remarks",
                ],
                ["1", "Sun room work", "1", "40000", "40000", ""],
                ["2", "Chockat fixing", "1", "2500", "2500", ""],
                ["Total", "", "", "", "42500", ""],
                ["Materials", "", "", "", "", ""],
                ["1", "Blocks", "1000", "72", "72000", ""],
                ["2", "Stone", "3", "4000", "12000", ""],
                ["Total", "", "", "", "84000", ""],
                ["Miscellaneous & Administration", "", "", "", "", ""],
                ["1", "Staff food", "", "", "2100", ""],
                ["", "Nazam Chokidar Salary Jun 2026", "", "", "", ""],
                ["2", "Bike petrol", "", "", "500", ""],
                ["Total", "", "", "", "2600", ""],
                ["Total Amount", "", "", "", "129100", ""],
                ["Expense paid last req. 15 till", "", "", "", "9546797", ""],
            ]
            df = pd.DataFrame(sample_data)
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, header=False)
            output.seek(0)
            st.download_button(
                label="📥 Download Template",
                data=output,
                file_name="Sample_Requisition_Template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_template",
            )

        uploaded_file = st.file_uploader(
            "Choose Excel file",
            type=["xlsx", "xls"],
            help="Upload any requisition Excel file - works with any format!",
        )

        if uploaded_file:
            # Preview the data
            if st.button("👁️ Preview Data", use_container_width=True):
                try:
                    df = pd.read_excel(
                        uploaded_file, header=None, engine="openpyxl"
                    )
                    st.dataframe(df, use_container_width=True)
                except Exception as e:
                    st.error(f"Error previewing file: {e}")

            if st.button(
                "📥 Import Excel Data",
                type="primary",
                use_container_width=True,
            ):
                with st.spinner("Importing data..."):
                    # Use the smart importer
                    success, result, count = import_excel_data(
                        uploaded_file, user["id"]
                    )
                    if success:
                        st.success(f"✅ {result}")
                        st.balloons()
                        st.rerun()
                    else:
                        st.error(f"❌ {result}")

        st.markdown("</div>", unsafe_allow_html=True)

    # ------------------------------------------------------------------
    # Export Reports section
    # ------------------------------------------------------------------
    with st.expander("📄 Export Reports", expanded=False):
        st.info("Generate PDF reports for requisitions")
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id, name FROM projects ORDER BY name")
        projects = cursor.fetchall()
        conn.close()

        if projects:
            project_names = [p[1] for p in projects]
            selected_project = st.selectbox("Select Project", project_names)
            if st.button(
                "📊 Export Project Summary", use_container_width=True
            ):
                conn = get_db_connection()
                cursor = conn.cursor()
                cursor.execute(
                    "SELECT id FROM projects WHERE name = ?",
                    (selected_project,),
                )
                project = cursor.fetchone()
                conn.close()
                if project:
                    from utils.pdf_generator import generate_project_summary_pdf

                    generate_project_summary_pdf(project["id"])
        else:
            st.warning("No projects available")

    # ------------------------------------------------------------------
    # Filter & Create New Requisition
    # ------------------------------------------------------------------
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, name FROM projects ORDER BY name")
    projects = cursor.fetchall()
    conn.close()

    project_options = {p[1]: p[0] for p in projects}
    project_options["All Projects"] = None

    col1, col2 = st.columns([2, 1])
    with col1:
        selected_project_name = st.selectbox(
            "Filter by Project",
            list(project_options.keys()),
            key="req_project_filter",
        )
        project_id = project_options[selected_project_name]

    with col2:
        status_filter = st.selectbox(
            "Filter by Status",
            [
                "All",
                "DRAFT",
                "SUBMITTED",
                "VERIFIED",
                "APPROVED",
                "REJECTED",
                "RETURNED",
            ],
        )

    if user["role"] in ["ADMIN", "DATA_ENTRY"] and projects:
        with st.expander("➕ Create New Requisition", expanded=False):
            st.markdown("### Requisition Details")
            with st.form("create_requisition_form"):
                col1, col2, col3 = st.columns(3)

                with col1:
                    project_name = st.selectbox(
                        "Select Project*", [p[1] for p in projects]
                    )

                    conn = get_db_connection()
                    cursor = conn.cursor()
                    cursor.execute(
                        "SELECT opening_balance, current_balance, code FROM projects WHERE name = ?",
                        (project_name,),
                    )
                    proj = cursor.fetchone()
                    conn.close()

                    if proj:
                        opening_balance = (
                            proj["opening_balance"]
                            if proj["opening_balance"] is not None
                            else 0
                        )
                        current_balance = (
                            proj["current_balance"]
                            if proj["current_balance"] is not None
                            else 0
                        )
                        st.info(f"Opening Balance: PKR {opening_balance:,.2f}")
                        st.info(f"Current Balance: PKR {current_balance:,.2f}")
                    else:
                        opening_balance = 0
                        current_balance = 0

                with col2:
                    conn = get_db_connection()
                    cursor = conn.cursor()
                    cursor.execute(
                        "SELECT code FROM projects WHERE name = ?",
                        (project_name,),
                    )
                    proj_code = cursor.fetchone()
                    conn.close()

                    project_code = proj_code["code"] if proj_code else "REQ"

                    ref_no = st.text_input(
                        "Reference No.*",
                        value=f"{project_code}-{datetime.now().strftime('%Y%m')}-{len(projects)+1:04d}",
                        help="Format: PROJECTCODE-YYYYMM-XXXX",
                    )
                    period_start = st.date_input("Period Start*")

                with col3:
                    conn = get_db_connection()
                    cursor = conn.cursor()
                    cursor.execute(
                        """
                        SELECT closing_balance FROM requisitions 
                        WHERE project_id = (SELECT id FROM projects WHERE name = ?)
                        AND status = 'APPROVED'
                        ORDER BY created_at DESC LIMIT 1
                    """,
                        (project_name,),
                    )
                    last_req = cursor.fetchone()
                    conn.close()

                    expense_paid_last = (
                        last_req["closing_balance"] if last_req else 0
                    )

                    st.text_input(
                        "Expense Paid Last Req",
                        value=f"PKR {expense_paid_last:,.2f}",
                        disabled=True,
                    )
                    period_end = st.date_input("Period End*")

                opening_balance_display = st.number_input(
                    "Opening Balance (Carried Forward)",
                    min_value=0.0,
                    step=1000.0,
                    value=float(expense_paid_last) if expense_paid_last else 0.0,
                    help="This is carried forward from the previous requisition",
                )

                submitted = st.form_submit_button(
                    "Create Requisition",
                    use_container_width=True,
                    type="primary",
                )

                if submitted and ref_no and period_start and period_end:
                    try:
                        conn = get_db_connection()
                        cursor = conn.cursor()
                        cursor.execute(
                            "SELECT id, opening_balance FROM projects WHERE name = ?",
                            (project_name,),
                        )
                        project = cursor.fetchone()

                        if project:
                            cursor.execute(
                                """
                                SELECT closing_balance FROM requisitions 
                                WHERE project_id = ? 
                                AND status = 'APPROVED'
                                ORDER BY created_at DESC LIMIT 1
                            """,
                                (project["id"],),
                            )
                            last_req = cursor.fetchone()
                            opening_balance = (
                                last_req["closing_balance"]
                                if last_req
                                else project["opening_balance"]
                            )

                            cursor.execute(
                                """
                                INSERT INTO requisitions 
                                (project_id, ref_no, period_start, period_end, opening_balance, 
                                 closing_balance, expense_paid_last_req, status, created_by_id)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """,
                                (
                                    project["id"],
                                    ref_no,
                                    period_start,
                                    period_end,
                                    opening_balance,
                                    opening_balance,
                                    expense_paid_last,
                                    "DRAFT",
                                    user["id"],
                                ),
                            )

                            req_id = cursor.lastrowid
                            conn.commit()
                            conn.close()

                            st.success(
                                f"✅ Requisition {ref_no} created successfully!"
                            )
                            st.info(
                                f"📋 Opening Balance: PKR {opening_balance:,.2f}"
                            )
                            st.info(
                                f"📋 Expense Paid Last Req: PKR {expense_paid_last:,.2f}"
                            )
                            st.session_state.editing_requisition_id = req_id
                            st.rerun()
                    except Exception as e:
                        st.error(f"Error creating requisition: {e}")

    # ------------------------------------------------------------------
    # Display Requisitions
    # ------------------------------------------------------------------
    conn = get_db_connection()
    cursor = conn.cursor()
    query = """
    SELECT r.*, p.name as project_name, u.full_name as created_by_name
    FROM requisitions r
    LEFT JOIN projects p ON r.project_id = p.id
    LEFT JOIN users u ON r.created_by_id = u.id
    WHERE 1=1
    """
    params = []
    if project_id:
        query += " AND r.project_id = ?"
        params.append(project_id)
    if status_filter != "All":
        query += " AND r.status = ?"
        params.append(status_filter)
    query += " ORDER BY r.created_at DESC"

    cursor.execute(query, params)
    requisitions = cursor.fetchall()
    conn.close()

    if requisitions:
        for req in requisitions:
            # IMPORTANT: Define req_id here for use in the financial summary
            req_id = req["id"]
            
            with st.container(border=True):
                ref_no = req["ref_no"] if req["ref_no"] is not None else "N/A"
                project_name = req["project_name"] if req["project_name"] is not None else "Unknown"
                total_amount = float(req["total_amount"]) if req["total_amount"] is not None else 0
                opening_balance = float(req["opening_balance"]) if req["opening_balance"] is not None else 0
                closing_balance = float(req["closing_balance"]) if req["closing_balance"] is not None else 0
                expense_paid_last = float(req["expense_paid_last_req"]) if req["expense_paid_last_req"] is not None else 0
                status = req["status"] if req["status"] is not None else "DRAFT"
                period_start = req["period_start"] if req["period_start"] is not None else ""
                period_end = req["period_end"] if req["period_end"] is not None else ""

                col1, col2, col3, col4, col5 = st.columns([2, 1.5, 1.2, 1, 0.8])

                with col1:
                    st.markdown(f"**Ref No: {ref_no}**")
                    st.markdown(f"📋 {project_name}")
                    st.markdown(f"📅 {period_start[:10] if period_start else ''} to {period_end[:10] if period_end else ''}")

                with col2:
                    st.markdown("**Amount**")
                    st.markdown(f"PKR {total_amount:,.2f}")
                    st.markdown(f"**Expense Paid Last:** PKR {expense_paid_last:,.2f}")

                with col3:
                    st.markdown(f"**Opening:** PKR {opening_balance:,.2f}")
                    st.markdown(f"**Closing:** PKR {closing_balance:,.2f}")

                with col4:
                    st.markdown("**Status**")
                    st.markdown(get_status_badge(status), unsafe_allow_html=True)

                with col5:
                    # Status action buttons
                    if status == "DRAFT" and user["role"] in ["ADMIN", "DATA_ENTRY"]:
                        if st.button("✏️ Edit", key=f"edit_req_{req_id}"):
                            st.session_state.editing_requisition_id = req_id
                            st.rerun()

                        if st.button("📤 Submit", key=f"submit_req_{req_id}"):
                            try:
                                conn = get_db_connection()
                                cursor = conn.cursor()
                                cursor.execute(
                                    """
                                    UPDATE requisitions 
                                    SET status = 'SUBMITTED', submitted_at = CURRENT_TIMESTAMP, submitted_by_id = ?
                                    WHERE id = ?
                                """,
                                    (user["id"], req_id),
                                )
                                conn.commit()
                                conn.close()
                                st.success("✅ Requisition submitted for verification!")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Error submitting: {e}")

                    # RETURNED status - Return to Draft and Edit
                    if status == "RETURNED" and user["role"] in ["ADMIN", "DATA_ENTRY"]:
                        if st.button("↩️ Return to Draft", key=f"return_to_draft_{req_id}"):
                            try:
                                conn = get_db_connection()
                                cursor = conn.cursor()
                                cursor.execute(
                                    """
                                    UPDATE requisitions 
                                    SET status = 'DRAFT'
                                    WHERE id = ?
                                """,
                                    (req_id,),
                                )
                                conn.commit()
                                conn.close()
                                st.success("✅ Requisition returned to DRAFT status. You can now edit it.")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Error returning to draft: {e}")
                        
                        if st.button("✏️ Edit", key=f"edit_returned_{req_id}"):
                            st.session_state.editing_requisition_id = req_id
                            st.rerun()

                    # SUBMITTED status - Return to Draft
                    if status == "SUBMITTED" and user["role"] in ["ADMIN", "DATA_ENTRY"]:
                        if st.button("↩️ Return to Draft", key=f"return_submitted_{req_id}"):
                            try:
                                conn = get_db_connection()
                                cursor = conn.cursor()
                                cursor.execute(
                                    """
                                    UPDATE requisitions 
                                    SET status = 'DRAFT'
                                    WHERE id = ?
                                """,
                                    (req_id,),
                                )
                                conn.commit()
                                conn.close()
                                st.success("✅ Requisition returned to DRAFT status.")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Error returning to draft: {e}")

                    if status == "SUBMITTED" and user["role"] in ["ADMIN", "VERIFIER"]:
                        if st.button("🔍 Verify", key=f"verify_req_{req_id}"):
                            st.session_state.verifying_requisition_id = req_id
                            st.rerun()

                    # VERIFIED status - Return to Draft
                    if status == "VERIFIED" and user["role"] in ["ADMIN", "DATA_ENTRY"]:
                        if st.button("↩️ Return to Draft", key=f"return_verified_{req_id}"):
                            try:
                                conn = get_db_connection()
                                cursor = conn.cursor()
                                cursor.execute(
                                    """
                                    UPDATE requisitions 
                                    SET status = 'DRAFT'
                                    WHERE id = ?
                                """,
                                    (req_id,),
                                )
                                conn.commit()
                                conn.close()
                                st.success("✅ Requisition returned to DRAFT status.")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Error returning to draft: {e}")

                    if status == "VERIFIED" and user["role"] in ["ADMIN", "CEO"]:
                        if st.button("✅ Approve", key=f"approve_req_{req_id}"):
                            st.session_state.approving_requisition_id = req_id
                            st.rerun()

                    # DELETE BUTTON - Only for DRAFT and RETURNED status
                    if status in ["DRAFT", "RETURNED"] and user["role"] in ["ADMIN", "DATA_ENTRY"]:
                        if st.button("🗑️ Delete", key=f"delete_req_{req_id}", type="primary"):
                            st.session_state.delete_requisition_id = req_id
                            st.rerun()

                    # PDF Button
                    if st.button("📄 PDF", key=f"pdf_req_{req_id}"):
                        with st.spinner("Generating PDF..."):
                            pdf_path, error = generate_requisition_pdf(req_id)
                            if pdf_path and os.path.exists(pdf_path):
                                with open(pdf_path, "rb") as f:
                                    st.download_button(
                                        label="📥 Download PDF",
                                        data=f,
                                        file_name=f"Requisition_{ref_no}.pdf",
                                        mime="application/pdf",
                                        key=f"download_pdf_{req_id}",
                                    )
                                os.remove(pdf_path)
                                st.success("✅ PDF generated successfully!")
                            else:
                                st.error(f"❌ PDF generation failed: {error}")

                # ================================================================
                # FINANCIAL SUMMARY - Now defined inside the loop with access to req_id
                # ================================================================
                # ================================================================
                # FINANCIAL SUMMARY - Corrected version with proper delta_color
                # ================================================================
                with st.expander("💰 Financial Summary", expanded=False):
                    st.markdown("### 📊 Project Financial Summary")
                    
                    # Get current financial data
                    conn = get_db_connection()
                    cursor = conn.cursor()
                    
                    # First check if the columns exist
                    cursor.execute("PRAGMA table_info(requisitions)")
                    columns = [col[1] for col in cursor.fetchall()]
                    
                    # Build query dynamically based on existing columns
                    select_fields = ["id", "total_amount", "opening_balance", "closing_balance"]
                    
                    # Add financial fields if they exist
                    fin_fields = ['paid_amount', 'rar_bills', 'retention_money', 'total_expense', 'profit_loss']
                    for field in fin_fields:
                        if field in columns:
                            select_fields.append(field)
                    
                    query = f"SELECT {', '.join(select_fields)} FROM requisitions WHERE id = ?"
                    cursor.execute(query, (req_id,))
                    fin_data = cursor.fetchone()
                    conn.close()
                    
                    if fin_data:
                        # Get values with defaults
                        total_amount = float(fin_data['total_amount'] or 0)
                        
                        # Safely get financial fields with defaults
                        paid_amount = float(fin_data['paid_amount']) if 'paid_amount' in fin_data.keys() and fin_data['paid_amount'] is not None else 0
                        rar_bills = float(fin_data['rar_bills']) if 'rar_bills' in fin_data.keys() and fin_data['rar_bills'] is not None else 0
                        retention_money = float(fin_data['retention_money']) if 'retention_money' in fin_data.keys() and fin_data['retention_money'] is not None else 0
                        total_expense = float(fin_data['total_expense']) if 'total_expense' in fin_data.keys() and fin_data['total_expense'] is not None else 0
                        profit_loss = float(fin_data['profit_loss']) if 'profit_loss' in fin_data.keys() and fin_data['profit_loss'] is not None else 0
                        opening_balance = float(fin_data['opening_balance'] or 0)
                        closing_balance = float(fin_data['closing_balance'] or 0)
                        
                        # Calculate derived values
                        balance = total_amount - paid_amount
                        total_received = rar_bills + retention_money
                        
                        # Create form for editing financial data
                        with st.form(f"financial_form_{req_id}"):
                            col1, col2, col3 = st.columns(3)
                            
                            with col1:
                                st.markdown("**Transaction Summary**")
                                st.text_input(
                                    "Total Amount",
                                    value=f"PKR {total_amount:,.2f}",
                                    disabled=True,
                                    key=f"total_amt_display_{req_id}"
                                )
                                
                                paid_amount_edit = st.number_input(
                                    "Paid Amount (PKR)*",
                                    min_value=0.0,
                                    step=1000.0,
                                    value=paid_amount,
                                    key=f"paid_amt_{req_id}",
                                    help="Enter the total paid amount for this project"
                                )
                                
                                # Calculate and display balance automatically
                                calculated_balance = total_amount - paid_amount_edit
                                st.text_input(
                                    "Balance",
                                    value=f"PKR {calculated_balance:,.2f}",
                                    disabled=True,
                                    key=f"balance_display_{req_id}"
                                )
                            
                            with col2:
                                st.markdown("**RAR & Retention**")
                                rar_bills_edit = st.number_input(
                                    "Amount Received in RAR (Bills)",
                                    min_value=0.0,
                                    step=1000.0,
                                    value=rar_bills,
                                    key=f"rar_bills_{req_id}",
                                    help="Total amount received in shape of RAR bills"
                                )
                                
                                retention_edit = st.number_input(
                                    "5% Retention Money",
                                    min_value=0.0,
                                    step=1000.0,
                                    value=retention_money,
                                    key=f"retention_{req_id}",
                                    help="5% retention money held back"
                                )
                                
                                # Calculate total received
                                calc_total_received = rar_bills_edit + retention_edit
                                st.text_input(
                                    "Total Received",
                                    value=f"PKR {calc_total_received:,.2f}",
                                    disabled=True,
                                    key=f"total_received_display_{req_id}"
                                )
                            
                            with col3:
                                st.markdown("**Expense & Profit**")
                                total_expense_edit = st.number_input(
                                    "Total Expense Made (PKR)",
                                    min_value=0.0,
                                    step=1000.0,
                                    value=total_expense,
                                    key=f"total_expense_{req_id}",
                                    help="Total expense made on this project"
                                )
                                
                                # Calculate profit automatically
                                calc_profit = calc_total_received - total_expense_edit
                                
                                # Display profit with color coding
                                profit_color = "#28a745" if calc_profit > 0 else "#dc3545"
                                profit_bg = "#d4edda" if calc_profit > 0 else "#f8d7da"
                                profit_text = "#155724" if calc_profit > 0 else "#721c24"
                                
                                st.markdown(f"""
                                <div style="background: {profit_bg}; 
                                            padding: 12px; 
                                            border-radius: 8px; 
                                            border: 2px solid {profit_color};
                                            margin-top: 5px;">
                                    <div style="font-size: 14px; color: #666; margin-bottom: 4px;">Profit/Loss</div>
                                    <div style="font-size: 24px; font-weight: bold; color: {profit_text};">
                                        PKR {calc_profit:,.2f}
                                    </div>
                                </div>
                                """, unsafe_allow_html=True)
                            
                            # Save button
                            col1, col2, col3 = st.columns([1, 2, 1])
                            with col2:
                                submitted = st.form_submit_button(
                                    "💾 Save Financial Summary",
                                    use_container_width=True,
                                    type="primary"
                                )
                            
                            if submitted:
                                try:
                                    # Calculate values
                                    new_balance = total_amount - paid_amount_edit
                                    new_total_received = rar_bills_edit + retention_edit
                                    new_profit = new_total_received - total_expense_edit
                                    
                                    # Update database
                                    conn = get_db_connection()
                                    cursor = conn.cursor()
                                    cursor.execute('''
                                        UPDATE requisitions 
                                        SET 
                                            paid_amount = ?,
                                            rar_bills = ?,
                                            retention_money = ?,
                                            total_expense = ?,
                                            profit_loss = ?
                                        WHERE id = ?
                                    ''', (
                                        paid_amount_edit,
                                        rar_bills_edit,
                                        retention_edit,
                                        total_expense_edit,
                                        new_profit,
                                        req_id
                                    ))
                                    conn.commit()
                                    conn.close()
                                    
                                    st.success("✅ Financial summary updated successfully!")
                                    st.balloons()
                                    st.rerun()
                                    
                                except Exception as e:
                                    st.error(f"Error updating financial summary: {e}")
                        
                        # Display current financial status in a clean format
                        st.markdown("---")
                        st.markdown("### 📊 Current Financial Status")
                        
                        col1, col2, col3, col4 = st.columns(4)
                        with col1:
                            st.metric("Total Amount", f"PKR {total_amount:,.2f}")
                        with col2:
                            st.metric("Paid Amount", f"PKR {paid_amount:,.2f}")
                        with col3:
                            st.metric("Balance", f"PKR {balance:,.2f}")
                        with col4:
                            # FIXED: Use 'normal' or 'inverse' for delta_color
                            # 'normal' shows green for positive, red for negative
                            # 'inverse' shows red for positive, green for negative
                            profit_delta_color = "normal" if profit_loss >= 0 else "inverse"
                            st.metric(
                                "Profit/Loss", 
                                f"PKR {profit_loss:,.2f}", 
                                delta_color=profit_delta_color
                            )
                        
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("RAR Bills", f"PKR {rar_bills:,.2f}")
                        with col2:
                            st.metric("5% Retention", f"PKR {retention_money:,.2f}")
                        with col3:
                            st.metric("Total Received", f"PKR {rar_bills + retention_money:,.2f}")
    else:
        st.info("No requisitions found. Create your first requisition!")


def show_requisition_editor(req_id):
    """Show the requisition editor interface"""
    user = st.session_state.auth["user"]
    conn = get_db_connection()
    cursor = conn.cursor()
    
    req = cursor.execute("SELECT * FROM requisitions WHERE id = ?", (req_id,)).fetchone()
    if not req:
        conn.close()
        st.error("Requisition not found")
        return
    
    project = cursor.execute("SELECT * FROM projects WHERE id = ?", (req['project_id'],)).fetchone()
    
    # Get categories - FIX: Check if status column exists
    cursor.execute("PRAGMA table_info(categories)")
    columns = [col[1] for col in cursor.fetchall()]
    has_status = 'status' in columns
    
    if has_status:
        cursor.execute("SELECT id, name FROM categories WHERE status = 'Active' ORDER BY name")
    else:
        cursor.execute("SELECT id, name FROM categories ORDER BY name")
    categories = cursor.fetchall()
    category_options = {c['name']: c['id'] for c in categories}
    
    cursor.execute('''
    SELECT t.*, c.name as category_name
    FROM transactions t
    LEFT JOIN categories c ON t.category_id = c.id
    WHERE t.requisition_id = ?
    ORDER BY t.sr_no
    ''', (req_id,))
    transactions = cursor.fetchall()
    conn.close()
    
    st.markdown(f"""
    <div style="background: #e3f2fd; padding: 15px; border-radius: 10px; margin-bottom: 20px;">
        <h3 style="margin: 0;">✏️ Editing: {req['ref_no']}</h3>
        <p style="margin: 5px 0 0 0; color: #666;">
            Project: {project['name'] if project else 'Unknown'} | 
            Period: {req['period_start'][:10] if req['period_start'] else ''} to {req['period_end'][:10] if req['period_end'] else ''} | 
            Opening: PKR {req['opening_balance']:,.2f}
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    if transactions:
        st.subheader("Current Items")
        
        # Group transactions by category
        categories_data = {}
        for trans in transactions:
            category = trans['category_name'] or "Uncategorized"
            if category not in categories_data:
                categories_data[category] = []
            categories_data[category].append({
                "Sr": trans['sr_no'],
                "Particulars": trans['particulars_raw'],
                "Qty": float(trans['qty']),
                "Unit": trans['unit'] or "",
                "Rate": float(trans['rate']),
                "Amount": float(trans['amount']),
                "Remarks": trans['remarks'] or ""
            })
        
        # Display each category with its items
        for category, items in categories_data.items():
            st.markdown(f"### 📂 {category}")
            df = pd.DataFrame(items)
            
            # Format numbers for display
            df_display = df.copy()
            df_display['Qty'] = df_display['Qty'].apply(lambda x: f"{x:,.2f}")
            df_display['Rate'] = df_display['Rate'].apply(lambda x: f"{x:,.2f}")
            df_display['Amount'] = df_display['Amount'].apply(lambda x: f"{x:,.2f}")
            
            st.dataframe(df_display, use_container_width=True, hide_index=True)
            
            # Category total
            category_total = sum(item['Amount'] for item in items)
            st.markdown(f"**Category Total:** PKR {category_total:,.2f}")
        
        # Grand total
        total = sum(t['amount'] for t in transactions)
        st.markdown(f"### 📊 Grand Total: PKR {total:,.2f}")
    
    st.subheader("Add New Item")
    with st.form("add_transaction_form"):
        col1, col2 = st.columns(2)
        with col1:
            # Category selection with option to add new
            category_names = list(category_options.keys())
            
            # Add option to create new category
            category_choice = st.selectbox(
                "Category*",
                category_names + ["➕ Add New Category..."],
                help="Select existing category or choose 'Add New Category' to create one"
            )
            
            # Show input for new category if selected
            new_category_name = None
            if category_choice == "➕ Add New Category...":
                new_category_name = st.text_input("New Category Name*", placeholder="Enter new category name")
                if new_category_name:
                    # Check if category already exists
                    conn = get_db_connection()
                    cursor = conn.cursor()
                    cursor.execute("SELECT id FROM categories WHERE name = ?", (new_category_name,))
                    existing = cursor.fetchone()
                    conn.close()
                    
                    if existing:
                        st.warning(f"⚠️ Category '{new_category_name}' already exists!")
                        new_category_name = None
                    else:
                        st.success(f"✅ New category '{new_category_name}' will be created")
            
            particulars = st.text_input("Particulars/Description*")
            
            # Add contractor selection
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT id, name FROM contractors WHERE status = 'Active' ORDER BY name")
            contractors = cursor.fetchall()
            conn.close()
            
            contractor_options = {"None": None}
            for c in contractors:
                contractor_options[c['name']] = c['id']
            
            selected_contractor_name = st.selectbox("Contractor (Optional)", list(contractor_options.keys()))
            contractor_id = contractor_options[selected_contractor_name]
            
        with col2:
            qty = st.number_input("Quantity", min_value=0.0, step=1.0, value=1.0)
            unit = st.text_input("Unit (e.g., nos, kg, ft)", "nos")
            rate = st.number_input("Rate (PKR)", min_value=0.0, step=10.0, value=0.0)
            is_lump_sum = st.checkbox("Lump Sum (No Qty×Rate)")
        
        remarks = st.text_area("Remarks")
        
        # SUBMIT BUTTON - FIX: Added back
        submitted = st.form_submit_button("➕ Add Item", use_container_width=True, type="primary")
        
        if submitted and particulars and category_choice:
            try:
                # Handle new category creation
                if category_choice == "➕ Add New Category...":
                    if not new_category_name:
                        st.error("Please enter a new category name")
                        st.stop()
                    category_name = new_category_name
                    
                    # Create new category
                    conn = get_db_connection()
                    cursor = conn.cursor()
                    cursor.execute(
                        "INSERT INTO categories (name, sort_order) VALUES (?, ?)",
                        (category_name, 999)
                    )
                    category_id = cursor.lastrowid
                    conn.commit()
                    conn.close()
                    st.success(f"✅ New category '{category_name}' created!")
                    
                    # Update category options
                    category_options[category_name] = category_id
                else:
                    category_name = category_choice
                    category_id = category_options[category_name]
                
                amount = rate if is_lump_sum else qty * rate
                
                conn = get_db_connection()
                cursor = conn.cursor()
                
                # Insert transaction with contractor_id
                cursor.execute('''
                INSERT INTO transactions 
                (requisition_id, category_id, master_item_id, particulars_raw, qty, unit, 
                 rate, amount, remarks, sr_no, is_lump_sum, entered_by_id, contractor_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (req_id, category_id, None, particulars, qty if not is_lump_sum else 0, 
                      unit, rate if not is_lump_sum else 0, amount, remarks, 
                      len(transactions) + 1, 1 if is_lump_sum else 0, user["id"], contractor_id))
                
                # Update requisition total
                cursor.execute("SELECT SUM(amount) FROM transactions WHERE requisition_id = ?", (req_id,))
                total = cursor.fetchone()[0] or 0
                
                # Get opening balance
                cursor.execute("SELECT opening_balance, expense_paid_last_req FROM requisitions WHERE id = ?", (req_id,))
                req_data = cursor.fetchone()
                opening_balance = req_data['opening_balance'] if req_data else 0
                expense_paid_last = req_data['expense_paid_last_req'] if req_data else 0
                
                # Calculate closing balance
                closing_balance = opening_balance + total
                
                cursor.execute('''
                UPDATE requisitions 
                SET total_amount = ?, closing_balance = ?, expense_paid_last_req = ?
                WHERE id = ?
                ''', (total, closing_balance, expense_paid_last, req_id))
                
                conn.commit()
                conn.close()
                
                st.success(f"✅ Item added successfully! New Total: PKR {total:,.2f}")
                st.balloons()
                st.rerun()
            except Exception as e:
                st.error(f"Error adding item: {e}")
    
    if transactions:
        st.subheader("Delete Item")
        trans_options = {f"{t['sr_no']}. {t['particulars_raw'][:50]}": t['id'] for t in transactions}
        trans_to_delete = st.selectbox("Select item to delete", list(trans_options.keys()))
        if st.button("🗑️ Delete Selected Item", type="primary"):
            try:
                trans_id = trans_options[trans_to_delete]
                conn = get_db_connection()
                cursor = conn.cursor()
                cursor.execute("DELETE FROM transactions WHERE id = ?", (trans_id,))
                
                cursor.execute("SELECT SUM(amount) FROM transactions WHERE requisition_id = ?", (req_id,))
                total = cursor.fetchone()[0] or 0
                
                cursor.execute("SELECT opening_balance FROM requisitions WHERE id = ?", (req_id,))
                req_data = cursor.fetchone()
                opening_balance = req_data['opening_balance'] if req_data else 0
                closing_balance = opening_balance + total
                
                cursor.execute('''
                UPDATE requisitions 
                SET total_amount = ?, closing_balance = ?
                WHERE id = ?
                ''', (total, closing_balance, req_id))
                
                conn.commit()
                conn.close()
                st.success("✅ Item deleted successfully!")
                st.rerun()
            except Exception as e:
                st.error(f"Error deleting item: {e}")
    
    if st.button("⬅️ Back to Requisitions"):
        st.session_state.editing_requisition_id = None
        st.rerun()


def import_excel_data(uploaded_file, user_id):
    """Wrapper function for Excel import"""
    try:
        # Save uploaded file temporarily in a cross-platform temp file
        suffix = os.path.splitext(uploaded_file.name)[1] or ".xlsx"
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
        try:
            tmp.write(uploaded_file.getbuffer())
            tmp.flush()
            tmp.close()

            # Import using the new importer
            importer = ExcelRequisitionImporter(tmp.name, user_id)
            success, message, count = importer.import_data()

            return success, message, count
        finally:
            # Clean up temp file
            try:
                os.remove(tmp.name)
            except Exception:
                pass
    except Exception as e:
        import traceback
        traceback.print_exc()
        return False, f"Error importing file: {str(e)}", 0